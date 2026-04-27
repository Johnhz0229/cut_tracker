import csv
import io
import json
import os
import secrets
from datetime import date as Date
from typing import Optional

import bcrypt
import openpyxl
import uvicorn
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import calculator
import database
import llm_client

load_dotenv()
database.init_db()

app = FastAPI(title="Fat Loss Tracker")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

ADMIN_SECRET = os.environ.get("ADMIN_SECRET", "")


def ok(data):
    return {"success": True, "data": data, "error": None}


def err(msg: str, status: int = 400):
    raise HTTPException(status_code=status, detail={"success": False, "data": None, "error": msg})


# ── Auth dependency ───────────────────────────────────────────────────────────

def get_current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(
            status_code=401,
            detail={"success": False, "data": None, "error": "Authorization required"},
        )
    token = authorization[7:]
    user = database.get_user_from_token(token)
    if not user:
        raise HTTPException(
            status_code=401,
            detail={"success": False, "data": None, "error": "Invalid or expired token"},
        )
    return user


# ── Auth routes ───────────────────────────────────────────────────────────────

class LoginIn(BaseModel):
    username: str
    password: str


class AdminCreateUserIn(BaseModel):
    username: str
    password: str
    admin_secret: str


@app.post("/api/login")
def login(body: LoginIn):
    user = database.get_user_by_username(body.username)
    if not user or not bcrypt.checkpw(
        body.password.encode(), user["password_hash"].encode()
    ):
        err("Invalid username or password", 401)
    token = secrets.token_hex(32)
    database.create_session(token, user["id"])
    return ok({"token": token, "username": user["username"]})


@app.post("/api/logout")
def logout(authorization: Optional[str] = Header(None)):
    if authorization and authorization.startswith("Bearer "):
        database.delete_session(authorization[7:])
    return ok({"logged_out": True})


def require_admin(x_admin_secret: Optional[str] = Header(None)):
    if not ADMIN_SECRET or x_admin_secret != ADMIN_SECRET:
        raise HTTPException(
            status_code=401,
            detail={"success": False, "data": None, "error": "Invalid admin secret"},
        )


class AdminVerifyIn(BaseModel):
    admin_secret: str


@app.post("/api/admin/verify")
def admin_verify(body: AdminVerifyIn):
    if not ADMIN_SECRET or body.admin_secret != ADMIN_SECRET:
        err("Invalid admin secret", 401)
    return ok({"verified": True})


@app.post("/api/admin/create-user")
def admin_create_user(body: AdminCreateUserIn):
    if not ADMIN_SECRET or body.admin_secret != ADMIN_SECRET:
        err("Invalid admin secret", 401)
    pw_hash = bcrypt.hashpw(body.password.encode(), bcrypt.gensalt()).decode()
    try:
        user = database.create_user(body.username, pw_hash)
    except ValueError as e:
        err(str(e), 409)
    token = secrets.token_hex(32)
    database.create_session(token, user["id"])
    return ok({"token": token, "username": user["username"]})


@app.get("/api/admin/users")
def admin_list_users(_: None = Depends(require_admin)):
    return ok(database.get_all_users())


@app.delete("/api/admin/users/{username}")
def admin_delete_user(username: str, _: None = Depends(require_admin)):
    deleted = database.delete_user_by_username(username)
    if not deleted:
        err(f"User '{username}' not found", 404)
    return ok({"deleted": username})


# ── Profile ───────────────────────────────────────────────────────────────────

class ProfileIn(BaseModel):
    height_cm: float
    weight_kg: float
    age: int
    sex: str
    protein_goal_g: float | None = None
    carbs_goal_g: float | None = None
    fat_goal_g: float | None = None


@app.post("/api/profile")
def create_profile(body: ProfileIn, user: dict = Depends(get_current_user)):
    if body.sex not in ("male", "female"):
        err("sex must be 'male' or 'female'")
    profile = database.upsert_profile(body.model_dump(), user["id"])
    return ok(profile)


@app.get("/api/profile")
def read_profile(user: dict = Depends(get_current_user)):
    profile = database.get_profile(user["id"])
    return ok(profile)


# ── Log ───────────────────────────────────────────────────────────────────────

class LogIn(BaseModel):
    date: str
    weight_kg: float
    food_description: str
    activity_description: str = ""
    exercise_description: str
    activity_multiplier: float = 1.2


@app.post("/api/log")
def log_day(body: LogIn, user: dict = Depends(get_current_user)):
    profile = database.get_profile(user["id"])
    if not profile:
        err("Profile not set up. Please create a profile first.", 422)

    cached = database.get_cached_llm_result(
        user["id"], body.food_description, body.exercise_description
    )
    if cached:
        protein_g = cached["protein_g"]
        carbs_g = cached["carbs_g"]
        fat_g = cached["fat_g"]
        calories_burned_exercise = cached["calories_burned_exercise"]
        llm_notes = cached["llm_notes"] or ""
        food_items_json = cached["food_items_json"] or "[]"
    else:
        try:
            llm_data = llm_client.analyze(
                body.food_description,
                body.activity_description,
                body.exercise_description,
                body.weight_kg,
                profile["age"],
                profile["sex"],
            )
        except TimeoutError:
            err("AI analysis timed out. Try again.", 503)
        except RuntimeError as e:
            err(str(e), 503)

        protein_g = llm_data["protein_g"]
        carbs_g = llm_data["carbs_g"]
        fat_g = llm_data["fat_g"]
        calories_burned_exercise = llm_data["calories_burned_exercise"]
        llm_notes = llm_data.get("notes", "")
        food_items_json = json.dumps(llm_data.get("food_items", []))

    tef = calculator.calculate_tef(protein_g, carbs_g, fat_g)
    calories_food = calculator.calculate_calories_food(protein_g, carbs_g, fat_g)
    bmr = calculator.calculate_bmr(
        body.weight_kg, profile["height_cm"], profile["age"], profile["sex"]
    )
    tdee = calculator.calculate_tdee(bmr, body.activity_multiplier, calories_burned_exercise)
    deficit = calculator.calculate_deficit(tdee, calories_food, tef["tef_total"])

    record = database.upsert_record({
        "date": body.date,
        "weight_kg": body.weight_kg,
        "food_description": body.food_description,
        "activity_description": body.activity_description,
        "exercise_description": body.exercise_description,
        "protein_g": protein_g,
        "carbs_g": carbs_g,
        "fat_g": fat_g,
        "calories_food": calories_food,
        **tef,
        "calories_burned_exercise": calories_burned_exercise,
        "activity_multiplier": body.activity_multiplier,
        "tdee": tdee,
        "deficit": deficit,
        "llm_notes": llm_notes,
        "food_items_json": food_items_json,
    }, user["id"])
    return ok(record)


# ── Records ───────────────────────────────────────────────────────────────────

@app.get("/api/records")
def list_records(days: int = 30, user: dict = Depends(get_current_user)):
    return ok(database.get_records(user["id"], days))


@app.get("/api/records/{date}")
def get_record(date: str, user: dict = Depends(get_current_user)):
    record = database.get_record_by_date(date, user["id"])
    if not record:
        err(f"No record found for {date}", 404)
    return ok(record)


@app.delete("/api/records/{date}")
def delete_record(date: str, user: dict = Depends(get_current_user)):
    deleted = database.delete_record(date, user["id"])
    if not deleted:
        err(f"No record found for {date}", 404)
    return ok({"deleted": date})


# ── Summary ───────────────────────────────────────────────────────────────────

@app.get("/api/summary")
def summary(user: dict = Depends(get_current_user)):
    return ok(database.get_summary(user["id"]))


# ── CSV Export ────────────────────────────────────────────────────────────────

CSV_COLUMNS = [
    "date", "weight_kg", "calories_food", "tef_total", "net_intake", "tdee", "deficit",
    "protein_g", "carbs_g", "fat_g", "calories_burned_exercise",
    "food_description", "activity_description", "exercise_description", "llm_notes",
]


@app.get("/api/export/csv")
def export_csv(days: int = 0, user: dict = Depends(get_current_user)):
    records = database.get_records(user["id"], days if days > 0 else None)

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(CSV_COLUMNS)

    for r in records:
        net_intake = round((r.get("calories_food") or 0) - (r.get("tef_total") or 0), 1)
        row = []
        for col in CSV_COLUMNS:
            if col == "net_intake":
                row.append(net_intake)
            else:
                val = r.get(col)
                if isinstance(val, float):
                    row.append(round(val, 1))
                else:
                    row.append(val if val is not None else "")
        writer.writerow(row)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=fat_loss_{Date.today()}.csv"},
    )


# ── Bulk Upload ──────────────────────────────────────────────────────────────

BULK_COLUMNS = ["date", "weight_kg", "food_description", "exercise_description", "activity_multiplier"]


def _parse_upload_file(file: UploadFile) -> list[dict]:
    """Parse a CSV or Excel file into a list of row dicts."""
    filename = (file.filename or "").lower()
    rows = []

    if filename.endswith(".csv"):
        content = file.file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            rows.append({
                (k or "").strip(): (v or "").strip()
                for k, v in row.items() if k
            })

    elif filename.endswith(".xlsx"):
        content = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({
                (headers[i] or "").strip(): (str(row[i]).strip() if row[i] is not None else "")
                for i in range(len(headers))
            })

    else:
        err("Unsupported file format. Please upload a .csv or .xlsx file.", 400)

    return rows


def _validate_and_normalize_rows(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Validate rows and separate valid from invalid. Returns (valid, errors)."""
    valid = []
    errors = []

    for i, row in enumerate(rows):
        row_num = i + 2  # 1-indexed + header row
        date = (row.get("date") or "").strip()
        weight_str = (row.get("weight_kg") or "").strip()
        food = (row.get("food_description") or "").strip()
        exercise = (row.get("exercise_description") or "").strip()
        mult_str = (row.get("activity_multiplier") or "").strip()

        # Validate required fields
        missing = []
        if not date:
            missing.append("date")
        if not weight_str:
            missing.append("weight_kg")
        if not food:
            missing.append("food_description")
        if not exercise:
            missing.append("exercise_description")

        if missing:
            errors.append({"row": row_num, "error": f"Missing required fields: {', '.join(missing)}"})
            continue

        try:
            weight = float(weight_str)
        except ValueError:
            errors.append({"row": row_num, "error": f"Invalid weight_kg: {weight_str!r}"})
            continue

        mult = 1.2
        if mult_str:
            try:
                mult = float(mult_str)
            except ValueError:
                errors.append({"row": row_num, "error": f"Invalid activity_multiplier: {mult_str!r}"})
                continue
            if mult < 1.0 or mult > 2.0:
                errors.append({"row": row_num, "error": f"activity_multiplier must be between 1.0 and 2.0"})
                continue

        valid.append({
            "row_num": row_num,
            "date": date,
            "weight_kg": weight,
            "food_description": food,
            "exercise_description": exercise,
            "activity_multiplier": mult,
        })

    return valid, errors


@app.post("/api/bulk-upload")
async def bulk_upload(file: UploadFile, user: dict = Depends(get_current_user)):
    profile = database.get_profile(user["id"])
    if not profile:
        err("Profile not set up. Please create a profile first.", 422)

    rows = _parse_upload_file(file)
    if not rows:
        err("The file is empty or has no data rows.", 400)

    valid_rows, row_errors = _validate_and_normalize_rows(rows)
    results = [{"row": e["row"], "date": "", "success": False, "error": e["error"]} for e in row_errors]

    for r in valid_rows:
        try:
            llm_data = llm_client.analyze(
                r["food_description"],
                "",
                r["exercise_description"],
                r["weight_kg"],
                profile["age"],
                profile["sex"],
            )
        except (TimeoutError, RuntimeError) as e:
            results.append({
                "row": r["row_num"], "date": r["date"],
                "success": False, "error": str(e),
            })
            continue

        protein_g = llm_data["protein_g"]
        carbs_g = llm_data["carbs_g"]
        fat_g = llm_data["fat_g"]
        calories_burned_exercise = llm_data["calories_burned_exercise"]
        llm_notes = llm_data.get("notes", "")
        food_items_json = json.dumps(llm_data.get("food_items", []))

        tef = calculator.calculate_tef(protein_g, carbs_g, fat_g)
        calories_food = calculator.calculate_calories_food(protein_g, carbs_g, fat_g)
        bmr = calculator.calculate_bmr(
            r["weight_kg"], profile["height_cm"], profile["age"], profile["sex"]
        )
        tdee = calculator.calculate_tdee(bmr, r["activity_multiplier"], calories_burned_exercise)
        deficit = calculator.calculate_deficit(tdee, calories_food, tef["tef_total"])

        database.upsert_record({
            "date": r["date"],
            "weight_kg": r["weight_kg"],
            "food_description": r["food_description"],
            "activity_description": "",
            "exercise_description": r["exercise_description"],
            "protein_g": protein_g,
            "carbs_g": carbs_g,
            "fat_g": fat_g,
            "calories_food": calories_food,
            **tef,
            "calories_burned_exercise": calories_burned_exercise,
            "activity_multiplier": r["activity_multiplier"],
            "tdee": tdee,
            "deficit": deficit,
            "llm_notes": llm_notes,
            "food_items_json": food_items_json,
        }, user["id"])

        results.append({
            "row": r["row_num"], "date": r["date"], "success": True,
        })

    successes = sum(1 for r in results if r["success"])
    failures = sum(1 for r in results if not r["success"])
    return ok({
        "total": len(results),
        "successes": successes,
        "failures": failures,
        "results": results,
    })


@app.get("/api/bulk-upload/template")
def bulk_upload_template(user: dict = Depends(get_current_user)):
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(BULK_COLUMNS)
    writer.writerow(["2026-04-27", "82.5", "3 eggs, 200g chicken, 100g rice", "45min weight training", "1.2"])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=bulk_upload_template.csv"},
    )


# ── Static frontend ───────────────────────────────────────────────────────────

@app.get("/admin", include_in_schema=False)
def admin_page():
    return FileResponse("frontend/admin.html")


@app.get("/manifest.json", include_in_schema=False)
def manifest():
    return FileResponse("frontend/manifest.json", media_type="application/manifest+json")


app.mount("/static", StaticFiles(directory="frontend/static"), name="static")
app.mount("/", StaticFiles(directory="frontend", html=True), name="frontend")


if __name__ == "__main__":
    port = int(os.getenv("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
